import json
import requests
import openpyxl
import re

myToken="YOUR_PERSONAL_ACCESS_TOKEN_HERE"
myRoom="YOUR_ROOM_ID_HERE"
myWorkbook = "lab05userlist.xlsx"
mySheetStartCol = 3
mySheetStartRow = 4
TestOnly = "yes"

def post_membership(mytoken,roomid,email,Moderator=False):
	# The header is send to authenticate
	header = {'Authorization':mytoken, 'content-type':'application/json'}
	# NEW: we have to include the email address and room moderator status
	payload = {'roomId':roomid,'personEmail':email,'isModerator':Moderator}
	# NEW: we want to create something: use POST instead of GET
	#Send POST request with the above header+payload.Put result in 'result'
	result = requests.post(url='https://api.ciscospark.com/v1/memberships',
						headers=header,json=payload)
	# Return POST request status, status '200'=succesful.
	returnData = json.loads(result.text)
	returnData['statuscode'] = str(result.status_code)
	return returnData


ListToAdd = list()
ListSuccess = list()
ListFailed = list()
# __ Open Spreadsheet
wb = openpyxl.load_workbook(myWorkbook)
# __ Getting name of FIRST sheet in workbook
mySheetName = wb.get_sheet_names()[0]
# __ Open the first sheet
sheet = wb.get_sheet_by_name(mySheetName)
# __ Get the highest row number (bottom of the sheet)
mySheetHighestRow = sheet.max_row + 1
# __ Print highest row
print ("highest row:", mySheetHighestRow)


# the range start-row to last-row, if that contains 200 cells,
# 'i' will go from 1 to 200
for i in range(mySheetStartRow, mySheetHighestRow, 1):
	# Put the cell value in row=i in variable "EmailFound"
	EmailFound = sheet.cell(row=i, column=mySheetStartCol).value
	# Check if the result is an email address ('@' followed by a '.')
	if not re.match(r"[^@]+@[^@]+\.[^@]+", EmailFound):
		# Add 'wrong' email addresses to a list: you can print them later
		print ("--- No valid email address found:", EmailFound)
		ListFailed.append(EmailFound)
		# Continue with "FOR" loop, skipping the 2 lines after continue
		continue
	if TestOnly.lower() == "yes": print (EmailFound)
	# The lines below only run if the email address is valid.
	ListToAdd.append(EmailFound)

if TestOnly.lower() == "no" :
	# loop through found email addressess
	for addname in ListToAdd :
		print ("====== ADD MEMBER:", addname)
		# Add email address to configured Spark room
		txt2 = post_membership("Bearer "+myToken,myRoom,addname)
		# Print results
		print ("TXT: " + str(txt2))
		# Store email address in list if an error occurs
		if txt2.get('statuscode') == "403" :
			print ("--- ERROR: status code is 403\n\n")
			ListFailed.append(addname)
		# Store email address in list if successfully added
		if txt2.get('statuscode') == "200" :
			ListSuccess.append(addname)

print ("\n\n ------------ Failed users: ------------ ")
# for each name found in the Failed List, print the name
for names in ListFailed :
	print(names)
# print the number of failed & success users (length of list)
print ("------------ failed count:", str(len(ListFailed)))
print ("------------ success count:", str(len(ListSuccess)))
